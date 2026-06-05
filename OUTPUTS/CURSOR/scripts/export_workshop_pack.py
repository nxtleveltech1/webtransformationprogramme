#!/usr/bin/env python3
"""Export WorkshopPack data to xlsx and regenerate executive HTML."""
from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
PACK = ROOT / "WorkshopPack"
CURSOR_OUT = ROOT / "OUTPUTS" / "CURSOR"
CURSOR_OUT.mkdir(parents=True, exist_ok=True)

# Agent deliverables live under OUTPUTS/CURSOR (primary). Also refresh canonical paths when writable.
XLSX_NAME = "Web_Transformation_Workshop Working Doc - 2.xlsx"
HTML_NAME = "portfolio-exec-2026-06-03.html"

def _xlsx_template() -> Path:
    for candidate in (
        PACK / "Web_Transformation_Workshop Working Doc - 2_export.xlsx",
        PACK / XLSX_NAME,
        PACK / "Web_Transformation_Workshop Working Doc - 2 (backup pre-AI 20260603-1001).xlsx",
    ):
        if candidate.exists():
            return candidate
    raise FileNotFoundError("No xlsx template found in WorkshopPack")


def _html_source() -> Path:
    p = ROOT / "Roadmap" / HTML_NAME
    if p.exists():
        return p
    p = CURSOR_OUT / HTML_NAME
    if p.exists():
        return p
    raise FileNotFoundError("No portfolio-exec HTML source found")

CURSOR_XLSX = CURSOR_OUT / XLSX_NAME
CURSOR_HTML = CURSOR_OUT / HTML_NAME
CANONICAL_XLSX = PACK / XLSX_NAME
CANONICAL_HTML = ROOT / "Roadmap" / HTML_NAME

sys.path.insert(0, str(Path(__file__).parent))
from workshop_data import (  # noqa: E402
    ACTIONS, CRITICAL_PATH, DECISIONS, DEPENDENCIES, GOVERNANCE,
    OVERVIEW_JOURNEYS, PROVISIONAL_NOTE, RAID, REGISTER_COUNTS,
    ROADMAP_ROWS, STREAM_INPUTS, TRADEOFFS,
)


def clear_data_rows(ws, start_row: int = 2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_sheet_data(ws, headers: list, rows: list, start_row: int = 2):
    clear_data_rows(ws, start_row)
    for i, row in enumerate(rows, start_row):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)


def export_xlsx(target: Path, template: Path | None = None):
    template = template or _xlsx_template()
    wb = openpyxl.load_workbook(template)
    # Revision note on Overview
    if "Overview" in wb.sheetnames:
        ws = wb["Overview"]
        ws["A1"] = PROVISIONAL_NOTE

    write_sheet_data(
        wb["Key Decisions"],
        [],
        DECISIONS,
    )
    write_sheet_data(
        wb["Trade off's"],
        [],
        TRADEOFFS,
    )
    act_rows = [
        (
            n,
            area,
            ws_name,
            pri,
            desc,
            owner,
            due,
            status,
            f"{notes} | {trace}" if notes else trace,
        )
        for n, area, ws_name, pri, desc, owner, due, status, notes, trace in ACTIONS
    ]
    write_sheet_data(wb["Action Log"], [], act_rows)

    raid_rows = []
    for row in RAID:
        if row[0] == "Risk":
            _, rid, desc, area, prob, impact, owner, mitigation, due, status = row
            raid_rows.append((row[0], rid, desc, area, prob, impact, owner, mitigation, due, status))
        else:
            _, rid, desc, impact, teams, owner, resolution, due, status = row
            raid_rows.append(
                ("Issue", rid, desc, teams, "—", impact, owner, resolution, due, status)
            )
    write_sheet_data(wb["RAID Log"], [], raid_rows)

    dep_rows = [
        (did, stream, dep, owner, req, impact, mit, esc, status)
        for did, stream, dep, owner, req, impact, mit, esc, status in DEPENDENCIES
    ]
    write_sheet_data(wb["Dependencies"], [], dep_rows)

    write_sheet_data(wb["Critical Path"], [], CRITICAL_PATH)
    write_sheet_data(wb["Roadmap"], [], ROADMAP_ROWS)
    write_sheet_data(wb["Governance"], [], GOVERNANCE)

    if "Overview" in wb.sheetnames:
        ws = wb["Overview"]
        for idx, (journey, ba, cluster, note) in enumerate(OVERVIEW_JOURNEYS, 2):
            ws.cell(row=idx, column=1, value=journey)
            ws.cell(row=idx, column=2, value=ba)
            ws.cell(row=idx, column=3, value=cluster)
            ws.cell(row=idx, column=8, value=note)

    if "Stream Inputs" in wb.sheetnames:
        write_sheet_data(
            wb["Stream Inputs"],
            [],
            STREAM_INPUTS,
        )

    if "Lead Prep Artefacts" in wb.sheetnames:
        ws = wb["Lead Prep Artefacts"]
        for r in range(2, min(ws.max_row, 20) + 1):
            status_cell = 8
            if ws.cell(row=r, column=1).value and str(ws.cell(row=r, column=1).value).startswith("Scope"):
                ws.cell(row=r, column=status_cell, value="Partial — Day 1 blockers surfaced")

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    print(f"Wrote xlsx: {target}")
    print(f"  Actions: {len(ACTIONS)}, RAID: {len(RAID)}, Dependencies: {len(DEPENDENCIES)}")


def tag_html(class_name: str, text: str) -> str:
    return f'<span class="tag {class_name}">{text}</span>'


def pri_tag(p: str) -> str:
    m = {"Critical": "critical", "High": "high", "Medium": "medium", "Low": "low"}
    return tag_html(m.get(p, "grey"), p)


def status_tag(s: str) -> str:
    s_lower = s.lower()
    if "progress" in s_lower:
        return tag_html("inprogress", s)
    if "suggest" in s_lower:
        return tag_html("grey", s)
    if "confirm" in s_lower:
        return tag_html("confirmed", s)
    if "open" in s_lower or "new" in s_lower:
        return tag_html("new", s)
    return tag_html("open", s)


def dec_status_tag(s: str) -> str:
    if "Confirmed" in s and "Proposed" not in s:
        return tag_html("confirmed", s.split("/")[0].strip())
    if "intent" in s.lower():
        return tag_html("intent", "Confirmed (intent)")
    return tag_html("proposed", s)


def impact_tag(impact: str) -> str:
    if impact == "High":
        return tag_html("high", impact)
    if impact == "Medium":
        return tag_html("medium", impact)
    return tag_html("low", impact)


def build_actions_table() -> str:
    rows = []
    for n, area, ws_name, pri, desc, owner, due, status, notes, trace in ACTIONS:
        act_id = f"ACT-{n:03d}"
        rows.append(
            f"<tr><td class=\"id\">{act_id}</td>"
            f"<td>{area}</td><td>{ws_name}</td>"
            f"<td>{pri_tag(pri)}</td>"
            f"<td>{desc}</td><td class=\"nowrap\">{owner}</td>"
            f"<td class=\"nowrap\">{due}</td><td>{status_tag(status)}</td>"
            f"<td style=\"font-size:11px;color:var(--muted)\">{notes}</td></tr>"
        )
    return "\n".join(rows)


def build_risks_table() -> str:
    rows = []
    for row in RAID:
        if row[0] != "Risk":
            continue
        _, rid, desc, area, prob, impact, owner, mitigation, due, status = row
        prob_cls = "high" if prob == "High" else "medium" if prob == "Medium" else "low"
        rows.append(
            f"<tr><td class=\"id\">{rid}</td><td>{desc}</td>"
            f"<td>{area}</td><td><span class=\"tag {prob_cls}\">{prob}</span></td>"
            f"<td>{impact_tag(impact)}</td><td class=\"nowrap\">{owner}</td>"
            f"<td style=\"font-size:11.5px\">{mitigation}</td></tr>"
        )
    return "\n".join(rows)


def build_issues_table() -> str:
    rows = []
    for row in RAID:
        if row[0] != "Issue":
            continue
        _, rid, desc, impact, teams, owner, resolution, due, status = row
        rows.append(
            f"<tr><td class=\"id\">{rid}</td><td>{desc}</td>"
            f"<td>{teams}</td><td>{impact_tag(impact)}</td>"
            f"<td class=\"nowrap\">{owner}</td>"
            f"<td style=\"font-size:11.5px\">{resolution}</td></tr>"
        )
    return "\n".join(rows)


def humanize_source(trace: str) -> str:
    """Turn internal trace codes (e.g. 'D2|S1|SRC-008|DEC-013') into an exec-friendly label."""
    day = sess = ""
    for part in trace.split("|"):
        p = part.strip()
        if re.fullmatch(r"D\d", p):
            day = "Day " + p[1]
        elif re.fullmatch(r"S\d", p):
            sess = "Session " + p[1]
    label = ", ".join(x for x in (day, sess) if x)
    return label or "Workshop"


def build_decisions_cards() -> str:
    cards = []
    for dec in DECISIONS:
        did, cat, item, desc, owner, req, status, due, trace = dec
        cls = "confirmed" if status == "Confirmed" or "Confirmed (intent)" in status else "proposed"
        cards.append(
            f'<div class="decision {cls}">'
            f'<div class="d-top"><span class="num">{did}</span>'
            f'<span class="title">{item}</span>{dec_status_tag(status)}</div>'
            f'<div class="body">{desc}</div>'
            f'<div class="meta"><strong>Owner:</strong> {owner} · {humanize_source(trace)}</div>'
            f"</div>"
        )
    return "\n".join(cards)


def patch_html(target: Path, source: Path | None = None):
    source = source or _html_source()
    if target == source:
        html = source.read_text(encoding="utf-8")
    else:
        html = source.read_text(encoding="utf-8")

    max_act = max(n for n, *_ in ACTIONS)
    max_rsk = max(int(r[1].split("-")[1]) for r in RAID if r[0] == "Risk")
    max_iss = max(int(r[1].split("-")[1]) for r in RAID if r[0] != "Risk")

    # Executive view: remove any internal/working callout banner entirely.
    banner_pat = r'<div class="callout warn[^>]*margin:0 56px 20px[^>]*>.*?</div>\s*'
    html = re.sub(banner_pat, "", html, count=1, flags=re.DOTALL)

    html = re.sub(
        r"The plan can only be baselined once the scope decision \(DEC-001\) is made\.",
        "Day 2 set the plan around phased, audience-based delivery (DEC-015):",
        html,
    )

    html = re.sub(
        r'<div class="src-note">.*?</div>',
        '<div class="src-note">Consolidated outcome of the two-day planning workshop — Day 1: 2 June 2026 · Day 2: 3 June 2026. Executive status as at 4 June 2026.</div>',
        html,
        count=1,
        flags=re.DOTALL,
    )

    # Register counts
    for key, val in [
        ("Decisions", REGISTER_COUNTS["decisions"]),
        ("Actions", REGISTER_COUNTS["actions"]),
        ("Risks", REGISTER_COUNTS["risks"]),
        ("Issues", REGISTER_COUNTS["issues"]),
        ("Assumptions", REGISTER_COUNTS["assumptions"]),
        ("Dependencies", REGISTER_COUNTS["dependencies"]),
        ("Open questions", REGISTER_COUNTS["questions"]),
        ("Parking lot", REGISTER_COUNTS["parking"]),
    ]:
        pass  # counts already correct in template

    # Replace decisions grid
    dec_pattern = r'(<section>\s*<h2>Decisions.*?</h2>\s*<div class="dec-grid">)(.*?)(</div>\s*</section>)'
    html = re.sub(
        dec_pattern,
        r"\1\n" + build_decisions_cards() + r"\n\3",
        html,
        count=1,
        flags=re.DOTALL,
    )

    # Insert full actions table after decisions section if not present
    actions_section = f"""
<section>
  <h2>Full action register <span class="soft">— {REGISTER_COUNTS['actions']} actions (ACT-001–{max_act:03d})</span></h2>
  <table class="tbl">
    <thead><tr><th>ID</th><th>Area</th><th>Workstream</th><th>Priority</th><th>Action</th><th>Owner</th><th>Due</th><th>Status</th><th>Related</th></tr></thead>
    <tbody>
{build_actions_table()}
    </tbody>
  </table>
</section>
"""
    if "Full action register" not in html:
        html = html.replace(
            "</main>\n\n<!-- ============================== PEOPLE",
            actions_section + "\n</main>\n\n<!-- ============================== PEOPLE",
            1,
        )
    else:
        html = re.sub(
            r'<section>\s*<h2>Full action register.*?</section>',
            actions_section.strip(),
            html,
            count=1,
            flags=re.DOTALL,
        )

    risks_section = f"""
<section>
  <h2>Full risk register <span class="soft">— {REGISTER_COUNTS['risks']} risks (RSK-001–{max_rsk:03d})</span></h2>
  <table class="tbl">
    <thead><tr><th>ID</th><th>Risk</th><th>Category</th><th>Prob</th><th>Impact</th><th>Owner</th><th>Mitigation</th></tr></thead>
    <tbody>
{build_risks_table()}
    </tbody>
  </table>
</section>
"""
    if "Full risk register" not in html:
        html = html.replace(
            actions_section.strip() + "\n</main>",
            actions_section.strip() + "\n" + risks_section + "\n</main>",
            1,
        )
    else:
        html = re.sub(
            r'<section>\s*<h2>Full risk register.*?</section>',
            risks_section.strip(),
            html,
            count=1,
            flags=re.DOTALL,
        )

    issues_section = f"""
<section>
  <h2>Live issues <span class="soft">— {REGISTER_COUNTS['issues']} issues (ISS-001–{max_iss:03d})</span></h2>
  <table class="tbl">
    <thead><tr><th>ID</th><th>Issue</th><th>Impact</th><th>Severity</th><th>Owner</th><th>Resolution</th></tr></thead>
    <tbody>
{build_issues_table()}
    </tbody>
  </table>
</section>
"""
    if "Live issues" not in html:
        html = html.replace(
            risks_section.strip() + "\n</main>",
            risks_section.strip() + "\n" + issues_section + "\n</main>",
            1,
        )
    else:
        html = re.sub(
            r'<section>\s*<h2>Live issues.*?</section>',
            issues_section.strip(),
            html,
            count=1,
            flags=re.DOTALL,
        )

    html = re.sub(
        r"<footer>.*?</footer>",
        f"<footer>\n  <strong>Old Mutual Web Transformation Programme</strong> — Executive status as at 4 June 2026.\n"
        f"  Two-day planning workshop · Cape Town · Facilitator: Gareth Bew. Confidential — for internal executive and steering distribution.\n"
        f"</footer>",
        html,
        count=1,
        flags=re.DOTALL,
    )

    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(html, encoding="utf-8")
    print(f"Patched HTML: {target}")


def mirror_workshop_pack():
    """Copy WorkshopPack markdown SSOT into OUTPUTS/CURSOR/WorkshopPack."""
    import shutil
    dest = CURSOR_OUT / "WorkshopPack"
    if dest.exists():
        shutil.rmtree(dest)
    shutil.copytree(
        PACK,
        dest,
        ignore=shutil.ignore_patterns("scripts", "*.xlsx", "*.xlsx.new"),
    )
    # Keep scripts at CURSOR root for re-export
    scripts_src = PACK / "scripts"
    scripts_dest = CURSOR_OUT / "scripts"
    if scripts_dest.exists():
        shutil.rmtree(scripts_dest)
    shutil.copytree(scripts_src, scripts_dest)
    print(f"Mirrored WorkshopPack -> {dest}")


def main():
    mirror_workshop_pack()
    export_xlsx(CURSOR_XLSX)
    patch_html(CURSOR_HTML)
    # Best-effort update canonical locations
    try:
        export_xlsx(CANONICAL_XLSX)
    except Exception as e:
        print(f"Note: canonical xlsx not updated ({e})")
    try:
        patch_html(CANONICAL_HTML, source=CURSOR_HTML)
    except Exception as e:
        print(f"Note: canonical HTML not updated ({e})")


if __name__ == "__main__":
    main()
